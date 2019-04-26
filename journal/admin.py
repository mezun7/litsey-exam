# coding=utf-8
from django.contrib import admin

# Register your models here.
from django.http import HttpResponse
from openpyxl.styles import Protection, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.writer.excel import save_virtual_workbook

from journal.models import Teacher, Class2, Subject, Mark, Student, MarksCoeff, Parallel


def export_xlsx(modeladmin, request, queryset):
    import openpyxl
    from django.utils.encoding import smart_str
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=student-info.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.get_active_sheet()
    ws.title = u"Информация об учениках"

    row_num = 0

    columns = [
        (u"ID", 15),
        (u"Фамилия", 20),
        (u"Имя", 20),
        (u"Отчество", 20),
        (u"Школа", 20),
        (u"Класс", 5),
        (u"Оплата питания", 15),
        (u"Справка со школы", 10),
        (u"Копия мед карты", 10),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
            obj.pk,
            smart_str(obj.fname),
            smart_str(obj.lname),
            smart_str(obj.fathers_name),
            smart_str(obj.school),
            smart_str(obj.class_name),
            obj.pay_for_eating,
            smart_str(obj.report_from_school),
            smart_str(obj.medical_card)

        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response


export_xlsx.short_description = u"Export XLSX"


def export_csv(StudentAdmin, request, queryset):
    import csv
    from django.utils.encoding import smart_str
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=student_info.csv'
    writer = csv.writer(response, csv.excel)
    response.write(u'\ufeff'.encode('utf8'))  # BOM (optional...Excel needs it to open UTF-8 file properly)
    writer.writerow([
        smart_str(u"ID"),
        smart_str(u"Фамилия"),
        smart_str(u"Имя"),
        smart_str(u"Отчество"),
        smart_str(u'Школа'),
        smart_str(u"Класс"),
        smart_str(u'Оплата питания'),
        smart_str(u"Справка со школы"),
        smart_str(u"Копия мед карты"),
    ])
    for obj in queryset:
        writer.writerow([
            smart_str(obj.pk),
            smart_str(obj.fname),
            smart_str(obj.lname),
            smart_str(obj.fathers_name),
            smart_str(obj.school),
            smart_str(obj.class_name),
            smart_str(obj.pay_for_eating),
            smart_str(obj.report_from_school),
            smart_str(obj.medical_card),
        ])
    return response


export_csv.short_description = u"Export CSV"


def export_xls_zip_all_classes(StudentAdmin, request, queryset):
    import openpyxl
    from django.utils.encoding import smart_str
    from openpyxl.utils import get_column_letter
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=journal.xlsx'
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    pos = 1

    columns = [
        (u"ID", 15),
        (u"ФИО", 20),
        (u"Класс", 20),
        (u"Оценка", 10),
        (u"Комментарий", 40),
    ]
    dv = DataValidation(type="decimal",
                        operator="between",
                        formula1=0,
                        formula2=10,
                        allow_blank=True)

    for parallel in Parallel.objects.all():
        for group in parallel.class2_set.all():
            isinstance(group, Class2)
            ws = wb.create_sheet(str(parallel) + str(group.name), pos)
            ws.protection.sheet = True
            ws.protection.password = 'litsey7'
            pos += 1
            row_num = 0
            for col_num in xrange(len(columns)):
                c = ws.cell(row=row_num + 1, column=col_num + 1)
                c.value = columns[col_num][0]

                #c.style.font.bold = True
                # set column width
                ws.column_dimensions[get_column_letter(col_num + 1)].width = columns[col_num][1]
            for student in group.student_set.all():

                row_num += 1
                row = [
                    student.pk,
                    smart_str(student.fname + " " + student.lname + ' ' + student.fathers_name),
                    smart_str(str(parallel.name) + str(group.name)),
                    smart_str(''),
                    smart_str(''),
                ]
                for col_num in xrange(len(row)):

                    c = ws.cell(row=row_num + 1, column=col_num + 1)
                    c.alignment = Alignment(wrap_text=True)
                    lock = False
                    if col_num + 1 <= 3:
                        lock = True
                    c.protection = Protection(locked=lock)
                    c.value = row[col_num]

    response.write(save_virtual_workbook(wb))
    return response


export_xls_zip_all_classes.short_descrption = u'Export to ZIP all Classes'


class TeacherAdmin(admin.ModelAdmin):
    list_display = ('user', 'subject')
    search_fields = ('user', 'subject')


class ClassAdmin(admin.ModelAdmin):
    list_display = ('parallel', 'name', 'class1_teacher')
    search_fields = ('name',)


class StudentAdmin(admin.ModelAdmin):
    actions = [export_csv, export_xlsx, export_xls_zip_all_classes]
    list_display = ('fname', 'lname', 'fathers_name', 'school', 'class_name')
    search_fields = ('lname', 'fname', 'school')


class MarkAdmin(admin.ModelAdmin):
    list_display = ('student', 'teacher', 'date', 'mark')
    search_fields = ('student', 'teacher', 'mark')
    list_filter = ('date',)


class ParallelAdmin(admin.ModelAdmin):
    list_display = ('name',)


admin.site.register(Teacher, TeacherAdmin)
admin.site.register(Class2, ClassAdmin)
admin.site.register(Subject)
admin.site.register(Student, StudentAdmin)
admin.site.register(Mark, MarkAdmin)
admin.site.register(MarksCoeff)
admin.site.register(Parallel, ParallelAdmin)
